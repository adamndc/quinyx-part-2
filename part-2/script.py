import csv
import datetime
from os import close
from openpyxl import load_workbook
from openpyxl.descriptors.base import DateTime #used instead of xlrd as it no longer supports .xlsx files

#load in workbook
workbook = load_workbook('sales data 2019 1-3.xlsx')
#grab first sheet
sheet = workbook.active

#Sheet => list format with date and sale information
data = []
for row in sheet.iter_rows(min_row=2):
    new_row =[]
    new_row.append((datetime.datetime.combine(row[2].value, row[3].value)))
    new_row.append((row[5].value))
    data.append(new_row)

#for storing new data
new_data = []

#for creating time buckets
TIME_DELTA = datetime.timedelta(minutes=15)
START_DATE = datetime.datetime.combine(data[1][0].date(), datetime.time(hour=10)) 
END_DATE = datetime.datetime.combine(data[-1][0].date() + datetime.timedelta(days=1), datetime.time(0))


CLOSED = datetime.time(hour=22, minute=30) #found through analysis in excel
TO_OPEN = datetime.timedelta(hours=9, minutes=30) #found through analysis in excel

current = START_DATE
while current < END_DATE:
    if current.time() >= CLOSED:
        current = current + TO_OPEN
    to_add = []
    to_add.append(current)
    to_add.append(0)
    to_add.append(0)
    new_data.append(to_add)
    current = current + TIME_DELTA

current_bucket = 0
current_data_row = 0 

while current_bucket < len(new_data) and current_data_row < len(data):
    if new_data[current_bucket + 1][0] < data[current_data_row][0]:
        current_bucket += 1
    else:
        new_data[current_bucket][1] += data[current_data_row][1]
        new_data[current_bucket][2] += 1
        current_data_row += 1    

with open('results.csv', 'w') as target:
    csv_writer = csv.writer(target)
    csv_writer.writerow(['Time', 'Sales', 'Transactions'])
    for row in new_data:
        csv_writer.writerow(row)